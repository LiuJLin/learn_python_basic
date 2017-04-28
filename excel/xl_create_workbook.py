from openpyxl import Workbook
#创建一个工作簿
wb = Workbook()
#获取创作的工作簿里的工作表
ws = wb.active
#创建新的工作表
ws1 = wb.create_sheet("Mysheet")
ws2 = wb.create_sheet("Mysheet",0)
#用title属性改变工作簿名称
ws.title = "New Title"
#更改选项卡的背景颜色
ws.sheet_properties.tabColor = "1072BA"
#可以将工作表的名字作为键
ws3 = wb["New Title"]
#打印工作簿的所有工作表名称
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

#只能在同一个工作簿下，创建工作表副本,不可以在不同工作簿之间。只能复制单元格和样式
source = wb.active
target = wb.copy_worksheet(source)
print(wb.sheetnames)
