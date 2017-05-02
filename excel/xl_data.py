import openpyxl
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl import cell
import datetime
#创建一个工作簿
wb =  openpyxl.load_workbook('test.xlsx')
#获取创作的工作簿里的工作表
ws = wb.active
#创建新的工作表
ws1 = wb.create_sheet("Mysheet")
ws2 = wb.create_sheet("Mysheet",0)
#用title属性改变工作簿名称
ws.title = "New Title"
#更改选项卡的背景颜色
ws.sheet_properties.tabColor = "1072BA"
#可以将工作表的名字作为键
ws3 = wb["New Title"]
#打印工作簿的所有工作表名称
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

c = ws['A4']

d = ws.cell(row=4, column=2, value=10)

for i in range(1,101):
    for j in range(1,101):
        ws.cell(row = i, column = j)

# 数据存储
c.value = "hello, world"
print(c,": ",c.value)
d.value = 3.14
print(d.value)

#wb = Workbook(guess_types = True)
c.value = '12%'
print(c.value)

##
d.value = datetime.datetime.now()
print(d.value)
c.value = '31.50'
print(c.value)
