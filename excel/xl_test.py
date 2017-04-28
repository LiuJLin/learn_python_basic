import openpyxl
from openpyxl.utils import get_column_letter

# 从当前目录读取test.xlsx
wb = openpyxl.load_workbook('test.xlsx')
print("Type of wb:\n",type(wb))
#获取所有sheet的名单
sheet_names = wb.get_sheet_names()
print("All sheets' names:\n",sheet_names)
#自由选取一个sheet
sheet = wb.get_sheet_by_name('Sheet1')
print("I choose the sheet:",sheet)
print("Type of sheet:\n",type(sheet))
#获取活动工作簿，即你当前所处的，或者是上次关闭前所处的
act_sheet = wb.get_active_sheet()
print("The active sheet is: ",act_sheet)

# 从表中获取单元格内容
cell_A1 = sheet['A1']
print("Content in A1 is:\n",cell_A1.value)

#row,column,coordinate 都没有，为什么
#print("Row " +　str(cell_B1.row) + ", Column " + cell_B1.column + " is " + cell_B1.value)
#print("Cell" + cell_B1.coordinate + " is " + cell_B1.value)

cell2 = sheet.cell(row = 1, column = 2)
print(cell2, "'s value is ", cell2.value )

for i in range(1, 8, 2):
    print(i, cell2.value)

rows = sheet.get_highest_row()
column = sheet.get_highest_column()
print("Rows :",rows)
print("Column:",column)
